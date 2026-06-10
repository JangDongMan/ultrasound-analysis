"""
초음파 ADC 신호를 Excel로 내보내기
- 원본 신호, 정류화(|ADC-128|), 엔벨로프(이동평균) 포함
- 엔벨로프 공식 = AVERAGE(INDEX 이동평균) -- 엑셀 수식으로 삽입

사용법:
    python3 export_to_excel.py <csv파일경로> [윈도우크기=30]

예시:
    python3 export_to_excel.py "usdata/data/260225/박성수/박성수_20260225_132408_(01)이마 중앙_M.csv"
    python3 export_to_excel.py path/to/file.csv 50
"""

import sys
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ───── 설정 ─────────────────────────────────────────────────────────────
SAMPLE_INTERVAL_NS  = 10        # 100MHz ADC → 10ns/sample
DROP_SAMPLES        = 1200      # 저장 시작 절대시간 오프셋
SPEED_OF_SOUND      = 1540.0    # mm/μs → m/s
THRESHOLD           = 100       # 엔벨로프 클러스터 임계값
DEFAULT_WINDOW      = 8         # 이동평균 반폭 (±행)
# ────────────────────────────────────────────────────────────────────────


def load_csv(path: str) -> np.ndarray:
    data = []
    with open(path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    data.append(int(line))
                except ValueError:
                    pass
    return np.array(data, dtype=np.float64)


def time_us(sample_idx: int) -> float:
    """샘플 인덱스 → 절대 시간(μs)"""
    return (DROP_SAMPLES + sample_idx) * SAMPLE_INTERVAL_NS / 1000.0


def export(csv_path: str, window: int = DEFAULT_WINDOW):
    adc = load_csv(csv_path)
    n   = len(adc)

    # 출력 파일명: 같은 디렉터리에 .xlsx
    base     = os.path.splitext(csv_path)[0]
    out_path = base + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Signal"

    # ── 파라미터 셀 ──────────────────────────────────────────────
    # H1 = 윈도우 반폭 (수식에서 $H$1 로 참조)
    ws["G1"] = "윈도우(±)"
    ws["G1"].font = Font(bold=True, color="FFFF00")
    ws["H1"] = window
    ws["H1"].font = Font(bold=True, color="FFFF00")

    # ── 헤더 행 (2행) ────────────────────────────────────────────
    headers = ["샘플#", "시간(μs)", "ADC 원본", "|ADC-128| 정류화", "엔벨로프(이동평균)", "임계값(100)"]
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── 데이터 + 수식 삽입 (3행~) ────────────────────────────────
    DATA_START = 3   # 데이터 시작 행
    for i, v in enumerate(adc):
        row = DATA_START + i
        t   = time_us(i)

        # A: 샘플 번호
        ws.cell(row=row, column=1, value=i)
        # B: 시간(μs)
        ws.cell(row=row, column=2, value=round(t, 3))
        # C: ADC 원본
        ws.cell(row=row, column=3, value=int(v))
        # D: |ADC-128| 정류화  → 엑셀 수식
        ws.cell(row=row, column=4, value=f"=ABS(C{row}-128)")
        # E: 엔벨로프 이동평균 (사용자 제시 공식 변형)
        #    =AVERAGE(INDEX(D:D, MAX(데이터시작행, ROW()-$H$1))
        #            :INDEX(D:D, MIN(COUNTA(D:D)+데이터시작행-1, ROW()+$H$1)))
        #   COUNTA(D:D) = 헤더(1) + 데이터(n) = n+2 → 마지막 데이터행 = n+2
        ws.cell(row=row, column=5,
                value=(f"=AVERAGE("
                       f"INDEX(D:D,MAX({DATA_START},ROW()-$H$1))"
                       f":"
                       f"INDEX(D:D,MIN(COUNTA(D:D)+{DATA_START-2},ROW()+$H$1))"
                       f")"))
        # F: 임계값 100 (수평 기준선)
        ws.cell(row=row, column=6, value=THRESHOLD)

    # ── 열 너비 ──────────────────────────────────────────────────
    widths = [8, 10, 10, 18, 22, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── 차트 ─────────────────────────────────────────────────────
    chart = LineChart()
    chart.title   = os.path.basename(csv_path)
    chart.style   = 10
    chart.y_axis.title = "Amplitude"
    chart.x_axis.title = "Time (μs)"
    chart.height  = 18
    chart.width   = 30

    # X축: 시간(μs) — B열
    x_vals = Reference(ws, min_col=2, min_row=DATA_START, max_row=DATA_START + n - 1)

    def make_series(col, title, color, width_emu, dash=None):
        ref = Reference(ws, min_col=col, min_row=DATA_START, max_row=DATA_START + n - 1)
        s = Series(ref, title=title)
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width     = width_emu
        if dash:
            s.graphicalProperties.line.dashDot = dash
        return s

    chart.series.append(make_series(3, "ADC 원본",           "888888", 5000))
    chart.series.append(make_series(4, "|ADC-128|",          "FFCC00", 8000))
    chart.series.append(make_series(5, f"엔벨로프(±{window})", "00E676", 14000))
    chart.series.append(make_series(6, "임계값(100)",         "FF4444", 8000, dash="dash"))

    chart.set_categories(x_vals)
    ws.add_chart(chart, "H3")

    wb.save(out_path)
    print(f"저장 완료: {out_path}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    csv_file   = sys.argv[1]
    win_size   = int(sys.argv[2]) if len(sys.argv) >= 3 else DEFAULT_WINDOW

    if not os.path.isfile(csv_file):
        print(f"파일 없음: {csv_file}")
        sys.exit(1)

    export(csv_file, win_size)
